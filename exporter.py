"""
exporter.py
-----------
Exports scraped results to a styled Excel file using openpyxl.
"""

from datetime import datetime
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from rich.console import Console

console = Console()

# Color palette
HEADER_BG = "1E3A5F"   # Dark navy
HEADER_FG = "FFFFFF"   # White
ALT_ROW_BG = "EBF3FB"  # Light blue
BORDER_COLOR = "B8CCE4"


def _make_border():
    thin = Side(style="thin", color=BORDER_COLOR)
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def export_to_excel(results: list[dict], output_path: str) -> str:
    """
    Export scraping results to a formatted Excel file.
    Each row = one website. Multiple emails/phones are joined with semicolons.

    Columns: #, Institution Name, Website, Email(s), Phone Number(s)
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Scraped Data"

    # ── Headers ──────────────────────────────────────────────────────────────
    headers = ["#", "Institution Name", "Type", "Website URL", "Email Address(es)", "Phone Number(s)"]
    header_font = Font(name="Calibri", bold=True, color=HEADER_FG, size=11)
    header_fill = PatternFill("solid", fgColor=HEADER_BG)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.row_dimensions[1].height = 30
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = _make_border()

    # ── Data Rows ─────────────────────────────────────────────────────────────
    alt_fill = PatternFill("solid", fgColor=ALT_ROW_BG)
    normal_font = Font(name="Calibri", size=10)
    wrap_align = Alignment(vertical="top", wrap_text=True)
    center_align = Alignment(horizontal="center", vertical="top")

    for row_idx, item in enumerate(results, start=2):
        ws.row_dimensions[row_idx].height = 40

        row_fill = alt_fill if row_idx % 2 == 0 else PatternFill()
        emails_str = " ; ".join(item.get("emails", [])) or "—"
        phones_str = " ; ".join(item.get("phones", [])) or "—"

        row_data = [
            row_idx - 1,
            item.get("name", "Unknown"),
            item.get("institute_type", "—"),
            item.get("url", ""),
            emails_str,
            phones_str,
        ]

        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = normal_font
            cell.border = _make_border()
            cell.fill = row_fill

            if col_idx == 1:
                cell.alignment = center_align
            else:
                cell.alignment = wrap_align

    # ── Column Widths ─────────────────────────────────────────────────────────
    col_widths = [5, 35, 18, 42, 42, 28]
    for col_idx, width in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # ── Auto Filter ───────────────────────────────────────────────────────────
    ws.auto_filter.ref = ws.dimensions

    # ── Freeze Top Row ────────────────────────────────────────────────────────
    ws.freeze_panes = "A2"

    # ── Summary Sheet ─────────────────────────────────────────────────────────
    ws_summary = wb.create_sheet("Summary")
    ws_summary["A1"] = "Education Institute Scraper — Report"
    ws_summary["A1"].font = Font(name="Calibri", bold=True, size=14, color=HEADER_BG)

    summary_data = [
        ("Total Institutions Found", len(results)),
        ("With Email", sum(1 for r in results if r.get("emails"))),
        ("With Phone", sum(1 for r in results if r.get("phones"))),
        ("Generated At", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
    ]
    for row_idx, (label, val) in enumerate(summary_data, start=3):
        ws_summary.cell(row_idx, 1, label).font = Font(name="Calibri", bold=True, size=10)
        ws_summary.cell(row_idx, 2, val).font = Font(name="Calibri", size=10)

    ws_summary.column_dimensions["A"].width = 30
    ws_summary.column_dimensions["B"].width = 25

    # ── Save ──────────────────────────────────────────────────────────────────
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))
    console.print(f"\n[green bold]Excel saved:[/green bold] [underline]{output_path}[/underline]")
    return str(output_path)
